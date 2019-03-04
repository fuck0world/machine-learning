# -*- coding: utf-8 -*-
from __future__ import print_function
import gzip
import os
import urllib
import json
import numpy 
import codecs
from sqlite3.test.userfunctions import func_isnone
from openpyxl import Workbook
from openpyxl import load_workbook

#DATA_FILE_NAME = '8wdata.xlsx'
DATA_FILE_NAME = 'expand_data.xlsx'

def validate_path(filename, work_directory):
    
    if not os.path.exists(work_directory):
      os.mkdir(work_directory)
    
    filepath = os.path.join(work_directory, filename)
    
    if not os.path.exists(filepath):
      return None
    return filepath


def choose_data(data_path, score_name, choose_num):
    
    final_DataFram = pd.DataFrame()
    for i in range(1,6):
        data = pd.read_csv(data_path)
        data = data.drop_duplicates(subset=None, keep='first' , inplace=False)
        data = data.loc[data[score_name] == i]
        data = data.reset_index(drop = True)
        row = data.iloc[:,0].size - 1
        #data = data.loc[np.random.randint(0, row, choose_num).tolist()]
        data = data.loc[range(0,choose_num)]
        
        final_DataFram = pd.concat([final_DataFram, data])
    final_DataFram = final_DataFram.reset_index(drop = True)
        
    return final_DataFram

def import_data(filename):

    
    wb = load_workbook(filename)
    
    ws = wb.get_sheet_by_name("Sheet1")
    row = ws.max_row
    col = ws.max_column
    
    arr1 = numpy.zeros( (row-1,col) )
    
    
    for row_A in range(2, row):
        for col_A in range(1, col + 1):
            arr1[row_A-2,col_A-1] = ws.cell(row = row_A, column = col_A).value
    return arr1

def dense_to_one_hot(labels_dense, num_classes):
    """Convert class labels from scalars to one-hot vectors."""
    num_labels = labels_dense.shape[0]
    index_offset = numpy.arange(num_labels) * num_classes
    labels_one_hot = numpy.zeros((num_labels, num_classes))
    labels_one_hot.flat[numpy.int_(index_offset + labels_dense.ravel())] = 1
    return labels_one_hot

def dense_to_one_hot_single(labels_dense, num_classes):
  """Convert class labels from scalars to one-hot vectors."""
  #index_offset = numpy.arange(1) * num_classes
  labels_one_hot = numpy.zeros((num_classes)).astype("float32")
  labels_one_hot.flat[numpy.int_(labels_dense)] = 1
  return labels_one_hot

class DataSet(object):

  def __init__(self, input, output, fake_data = False):
    if fake_data:
      self._num_examples = 10000
    else:
      assert input.shape[0] == output.shape[0], (
          "input.shape: %s output.shape: %s" % (input.shape,
                                                 output.shape))
      self._num_examples = input.shape[0]  
      
      input = input.astype(numpy.float32)
      output = output.astype(numpy.float32)
      
      
        
    self._input = input
    self._output = output
    self._epochs_completed = 0
    self._index_in_epoch = 0
    
  
  @property
  def input(self):
    return self._input

  @property
  def output(self):
    return self._output

  @property
  def num_examples(self):
    return self._num_examples

  @property
  def epochs_completed(self):
    return self._epochs_completed

  def next_batch(self, batch_size, fake_data=False):
    """Return the next `batch_size` examples from this data set."""
    if fake_data:
      fake_image = [1.0 for _ in range(784)]
      fake_label = 0
      return [fake_image for _ in range(batch_size)], [
          fake_label for _ in range(batch_size)]

    
    start = self._index_in_epoch
    self._index_in_epoch += batch_size
    if self._index_in_epoch > self._num_examples:
      # Finished epoch
      self._epochs_completed += 1
        
      # Shuffle the data
      
      perm = numpy.arange(self._num_examples)
      numpy.random.shuffle(perm)
      self._input = self._input[perm]
      self._output = self._output[perm]
        
      # Start next epoch
      start = 0
      self._index_in_epoch = batch_size
      assert batch_size <= self._num_examples
    end = self._index_in_epoch
    return self._input[start:end], self._output[start:end]

def read_data_sets(train_dir, fake_data=False):
  class DataSets(object):
    
    
    pass
  data_sets = DataSets()

  if fake_data:
    data_sets.train = DataSet([], [], fake_data=True)
    data_sets.validation = DataSet([], [], fake_data=True)
    data_sets.test = DataSet([], [], fake_data=True)
    return data_sets

  VALIDATION_SIZE = 1000
  TRAIN_SIZE = None

  local_file = validate_path(DATA_FILE_NAME, train_dir)
  if(local_file is None):
      print('Read File Error! The filepath does not exist!')
      exit()
  ori_data = import_data(local_file)
  ori_data = numpy.array(ori_data)
  
  print("max = ", numpy.max(ori_data[:,0:3]))
  print("min = ", numpy.min(ori_data[:,0:3]))
  
  
  #train_input = numpy.hstack( (ori_data[:TRAIN_SIZE,0:1]*1e-7, 
                               #ori_data[:TRAIN_SIZE,1:2]*1e-7, 
                               #ori_data[:TRAIN_SIZE,2:3]*1e-3, 
                               #ori_data[:TRAIN_SIZE,3:4]*1e-5, 
                               #dense_to_one_hot(ori_data[:TRAIN_SIZE,4],4),
                               #dense_to_one_hot(ori_data[:TRAIN_SIZE,5],2),
                               #dense_to_one_hot(ori_data[:TRAIN_SIZE,6],2)) )
    
  #test_input = numpy.hstack( (ori_data[TRAIN_SIZE:,0:1]*1e-7, 
                               #ori_data[TRAIN_SIZE:,1:2]*1e-7, 
                               #ori_data[TRAIN_SIZE:,2:3]*1e-3, 
                               #ori_data[TRAIN_SIZE:,3:4]*1e-5, 
                               #dense_to_one_hot(ori_data[TRAIN_SIZE:,4],4),
                               #dense_to_one_hot(ori_data[TRAIN_SIZE:,5],2),
                               #dense_to_one_hot(ori_data[TRAIN_SIZE:,6],2)) )
    
  train_input = numpy.hstack( (ori_data[:TRAIN_SIZE,0:1]*1e-4, 
                               ori_data[:TRAIN_SIZE,1:2]*1e-5, 
                               dense_to_one_hot(ori_data[:TRAIN_SIZE,2],2),
                               ori_data[:TRAIN_SIZE,2:3]*1e-4, 
                               ori_data[:TRAIN_SIZE,3:4]*1e-3,
                               ori_data[:TRAIN_SIZE,4:5]*1e-6, 
                               ori_data[:TRAIN_SIZE,5:6]*1e-4,
                               ori_data[:TRAIN_SIZE,6:7]*1e-4, 
                               ori_data[:TRAIN_SIZE,7:8]*1e-7,
                               ori_data[:TRAIN_SIZE,8:9]*1e-4, 
                               ori_data[:TRAIN_SIZE,9:10]*1e-5) )



  train_output = ori_data[:TRAIN_SIZE,[7,8,9,10]]*0.2
  
  #test_input = numpy.hstack( (ori_data[TRAIN_SIZE:,0:1]*1e-7, 
                               #ori_data[TRAIN_SIZE:,1:2]*1e-7, 
                               #ori_data[TRAIN_SIZE:,2:3]*1e-3, 
                               #ori_data[TRAIN_SIZE:,3:4]*1e-5, 
                               #dense_to_one_hot(ori_data[TRAIN_SIZE:,4],4),
                               #dense_to_one_hot(ori_data[TRAIN_SIZE:,5],2),
                               #dense_to_one_hot(ori_data[TRAIN_SIZE:,6],2)) )

  test_input = numpy.hstack( (ori_data[TRAIN_SIZE:,0:1]*1e-4, 
                               ori_data[TRAIN_SIZE:,1:2]*1e-5, 
                               dense_to_one_hot(ori_data[TRAIN_SIZE:,2],2),
                               ori_data[TRAIN_SIZE:,2:3]*1e-4, 
                               ori_data[TRAIN_SIZE:,3:4]*1e-3,
                               ori_data[TRAIN_SIZE:,4:5]*1e-6, 
                               ori_data[TRAIN_SIZE:,5:6]*1e-4,
                               ori_data[TRAIN_SIZE:,6:7]*1e-4, 
                               ori_data[TRAIN_SIZE:,7:8]*1e-7,
                               ori_data[TRAIN_SIZE:,8:9]*1e-4, 
                               ori_data[TRAIN_SIZE:,9:10]*1e-5) )
  

  test_output = ori_data[TRAIN_SIZE:,[7,8,9,10]]*0.2
  

  validation_input = test_input[:VALIDATION_SIZE]
  validation_output = test_output[:VALIDATION_SIZE]

  data_sets.train = DataSet(train_input, train_output)
  data_sets.validation = DataSet(validation_input, validation_output)
  data_sets.test = DataSet(test_input, test_output)

  return data_sets


if __name__ == '__main__':

    ddd=0
    aaa = import_data(filename = r'E:\\code\\python\\qoe_model\\expand_data.xlsx')
    bbb = read_data_sets(train_dir = r'E:\\code\\python\\qoe_model')
    ddd=0
    